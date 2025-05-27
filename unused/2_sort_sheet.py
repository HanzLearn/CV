import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def filter_and_group_by_frame(input_csv, output_excel):
    df = pd.read_csv(input_csv)

    df_rack1 = pd.DataFrame(columns=['frame_id', 'rack_id', 'track_id', 'class_id', 'class_name'])
    df_rack2 = pd.DataFrame(columns=['frame_id', 'rack_id', 'track_id', 'class_id', 'class_name'])
    df_rack3 = pd.DataFrame(columns=['frame_id', 'rack_id', 'track_id', 'class_id', 'class_name'])

    frame_ids = df['frame_id'].unique()

    for frame_id in frame_ids:
        frame_df = df[df['frame_id'] == frame_id]

        if 'Rack_1' in frame_df['rack_id'].values:
            rack1_df = frame_df[frame_df['rack_id'] == 'Rack_1']
            df_rack1 = pd.concat([df_rack1, rack1_df], ignore_index=True)

        if 'Rack_2' in frame_df['rack_id'].values:
            rack2_df = frame_df[frame_df['rack_id'] == 'Rack_2']
            df_rack2 = pd.concat([df_rack2, rack2_df], ignore_index=True)

        if 'Rack_3' in frame_df['rack_id'].values:
            rack3_df = frame_df[frame_df['rack_id'] == 'Rack_3']
            df_rack3 = pd.concat([df_rack3, rack3_df], ignore_index=True)

    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        df_rack1.to_excel(writer, sheet_name='Rack_1', index=False)
        df_rack2.to_excel(writer, sheet_name='Rack_2', index=False)
        df_rack3.to_excel(writer, sheet_name='Rack_3', index=False)

    def color_frame_id_changes(sheet_name):
        wb = load_workbook(output_excel)
        sheet = wb[sheet_name]

        fill_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        previous_frame_id = None
        for row in range(2, sheet.max_row + 1):
            current_frame_id = sheet.cell(row=row, column=1).value
            if current_frame_id != previous_frame_id:
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=row, column=col).fill = fill_color
            previous_frame_id = current_frame_id

        wb.save(output_excel)

    color_frame_id_changes('Rack_1')
    color_frame_id_changes('Rack_2')
    color_frame_id_changes('Rack_3')

    print(f"Data has been written to {output_excel} with color formatting")


def get_next_available_file(base_dir='received_data/CSV', output_dir='received_data/EXCEL'):
    counter = 1
    while True:
        input_file = os.path.join(base_dir, f"detection_results_{counter}.csv")
        output_file = os.path.join(output_dir, f"detection_results_colored_{counter}.xlsx")

        if not os.path.exists(input_file):
            return None

        if not os.path.exists(output_file):
            return input_file, output_file

        counter += 1


def process_files():
    os.makedirs('received_data/EXCEL', exist_ok=True)

    while True:
        file_pair = get_next_available_file()
        if not file_pair:
            print("No more files to process.")
            break

        input_csv, output_excel = file_pair
        print(f"Processing {input_csv}...")
        filter_and_group_by_frame(input_csv, output_excel)


if __name__ == "__main__":
    process_files()
