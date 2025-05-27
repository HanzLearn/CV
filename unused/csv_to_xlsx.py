import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

# Set up directories
csv_dir = 'received_data/CSV'
excel_dir = 'received_data/EXCEL'
os.makedirs(csv_dir, exist_ok=True)
os.makedirs(excel_dir, exist_ok=True)

# Get all detection_results*.csv files in received_data/CSV
csv_files = sorted([f for f in os.listdir(csv_dir) if re.match(r'detection_results(_\d+)?\.csv$', f)])

# Loop through all detection CSVs
for csv_file in csv_files:
    match = re.match(r'detection_results(_\d+)?\.csv$', csv_file)
    suffix = match.group(1) if match.group(1) else ''
    
    # File paths
    csv_path = os.path.join(csv_dir, csv_file)
    excel_file = f'colored_detections{suffix}.xlsx'
    excel_path = os.path.join(excel_dir, excel_file)

    # Skip if Excel version already exists
    if os.path.exists(excel_path):
        print(f"Skipping {csv_file} -> {excel_file} already exists.")
        continue

    print(f"Processing {csv_file} -> {excel_file}")

    # Step 1: Read the sorted CSV
    df = pd.read_csv(csv_path)

    # Step 2: Save to Excel format
    df.to_excel(excel_path, index=False)

    # Step 3: Open Excel file for formatting
    wb = load_workbook(excel_path)
    ws = wb.active

    # Define base frame colors
    frame_colors = [
        "FFFFCC",  # light yellow
        "CCFFFF",  # light cyan
        "FFCCCC",  # light red
        "CCE5FF",  # light blue
        "D5FFCC",  # light green
        "EBD6FF"   # light purple
    ]

    # Define two slightly different variants for rack toggling (same hue family)
    def adjust_shade(hex_color, lighten=True):
        factor = 1.1 if lighten else 0.9
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        r = min(255, int(r * factor))
        g = min(255, int(g * factor))
        b = min(255, int(b * factor))
        return f"{r:02X}{g:02X}{b:02X}"

    prev_frame = None
    prev_rack = None
    frame_color_idx = -1
    rack_toggle = False

    for row_idx in range(2, ws.max_row + 1):  # Skip header
        frame_id = ws.cell(row=row_idx, column=1).value
        rack_id = ws.cell(row=row_idx, column=2).value

        if frame_id != prev_frame:
            frame_color_idx = (frame_color_idx + 1) % len(frame_colors)
            prev_frame = frame_id
            prev_rack = None
            rack_toggle = False

        if rack_id != prev_rack:
            rack_toggle = not rack_toggle
            prev_rack = rack_id

        base_color = frame_colors[frame_color_idx].replace("#", "")
        final_color = adjust_shade(base_color, lighten=rack_toggle)
        fill = PatternFill(start_color=final_color, end_color=final_color, fill_type="solid")

        for col in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col).fill = fill

    wb.save(excel_path)
    wb.close()
    print(f"Saved {excel_file}")
